import openpyxl
import random
from datetime import date
from pathlib import Path  # <-- Added for finding your Downloads folder
from faker import Faker

# Initialize Faker
fake = Faker()

# Define structural choices matching the Django app validation limits
DEPARTMENTS = ["MATH", "SCIENCES", "LANGUAGES", "HUMANITIES", "TECHNICALS"]
PATHWAYS = ["STEM", "SOCIAL_SCIENCES", "ARTS_SPORTS", "GENERAL"]

# Predefined valid Subject names distributed across allowed departments
SUBJECT_TEMPLATES = [
    ("Mathematics", "MATH"), ("Advanced Calculus", "MATH"),
    ("Chemistry", "SCIENCES"), ("Biology", "SCIENCES"), ("Physics", "SCIENCES"),
    ("English Language", "LANGUAGES"), ("Kiswahili", "LANGUAGES"), ("French", "LANGUAGES"),
    ("History & Government", "HUMANITIES"), ("Geography", "HUMANITIES"),
    ("Computer Studies", "TECHNICALS"), ("Agriculture", "TECHNICALS")
]

# Predefined valid Classrooms matrix (Stream, Pathway, Classroom)
CLASSROOM_MATRIX = [
    ("Grade 10", "STEM", "East"),
    ("Grade 10", "STEM", "West"),
    ("Grade 10", "GENERAL", "Alpha"),
    ("Grade 10", "GENERAL", "Beta"),
    ("Grade 11", "STEM", "North"),
    ("Grade 11", "SOCIAL_SCIENCES", "Omega"),
    ("Grade 11", "GENERAL", "Gamma"),
    ("Grade 12", "STEM", "Delta"),
    ("Grade 12", "ARTS_SPORTS", "Sigma"),
    ("Grade 12", "GENERAL", "Theta"),
]

def generate_bulk_template():
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Instructions
    # ----------------------------------------------------
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr.append(["CBC School Data Import Template"])
    ws_instr.append([])
    ws_instr.append(["HOW TO USE THIS FILE"])
    ws_instr.append(["1. Fill in each sheet: Teachers, Classrooms, Subjects, Students."])
    ws_instr.append(["2. Row 1 is the header — do not edit or delete it."])
    ws_instr.append(["3. Row 2 in each sheet is an example/explanation row."])
    ws_instr.append(["4. Data rows start at Row 3 onwards."])
    ws_instr.append(["5. Re-running the command is safe: existing records are skipped, not duplicated."])
    
    # ----------------------------------------------------
    # Sheet 2: Subjects
    # ----------------------------------------------------
    ws_subs = wb.create_sheet(title="Subjects")
    ws_subs.append(["subject_name", "department"])
    ws_subs.append(["(e.g. Chemistry, English, Agriculture)", "(MATH | SCIENCES | LANGUAGES | HUMANITIES | TECHNICALS)"])
    
    for sub_name, dept in SUBJECT_TEMPLATES:
        ws_subs.append([sub_name, dept])
        
    # ----------------------------------------------------
    # Sheet 3: Teachers
    # ----------------------------------------------------
    ws_teachers = wb.create_sheet(title="Teachers")
    ws_teachers.append(["full_name", "username", "email", "tsc_number", "is_hod", "password"])
    ws_teachers.append([
        "(Full name as it appears officially)", "(Login username — no spaces)", 
        "(School email address (optional))", "(Leave blank if BOM teacher)", 
        "(Yes / No)", "(Temporary password)"
    ])
    
    for _ in range(40):
        full_name = fake.name()
        username = fake.unique.user_name().replace(".", "").replace("_", "")[:15]
        email = f"{username}@{fake.domain_name()}"
        tsc_number = f"TSC-{random.randint(100000, 999999)}" if random.random() > 0.2 else ""
        is_hod = "Yes" if random.random() > 0.85 else "No"
        password = "TemporaryPassword123!"
        
        ws_teachers.append([full_name, username, email, tsc_number, is_hod, password])

    # ----------------------------------------------------
    # Sheet 4: Classrooms
    # ----------------------------------------------------
    ws_rooms = wb.create_sheet(title="Classrooms")
    ws_rooms.append(["stream_name", "pathway", "classroom_name"])
    ws_rooms.append(["(e.g. Grade 10, Grade 11)", "(STEM | SOCIAL_SCIENCES | ARTS_SPORTS | GENERAL)", "(Section name)"])
    
    for stream, pathway, room in CLASSROOM_MATRIX:
        ws_rooms.append([stream, pathway, room])

    # ----------------------------------------------------
    # Sheet 5: Students
    # ----------------------------------------------------
    ws_students = wb.create_sheet(title="Students")
    ws_students.append(["admission_number", "name", "stream_name", "pathway", "classroom_name", "academic_year", "term", "date_of_birth"])
    ws_students.append(["(Unique school ID)", "(Full student name)", "(Stream Name)", "(Pathway)", "(Classroom Name)", "(Year)", "(Term 1-3)", "(YYYY-MM-DD)"])
    
    start_adm = 10001
    total_students = 750  
    
    for i in range(total_students):
        adm_number = str(start_adm + i)
        student_name = fake.name()
        stream, pathway, room = CLASSROOM_MATRIX[i % len(CLASSROOM_MATRIX)]
        academic_year = "2026"
        term = str(random.choice([1, 2, 3]))
        
        dob_object = date(random.randint(2008, 2012), random.randint(1, 12), random.randint(1, 28))
        
        if i % 2 == 0:
            date_of_birth = dob_object.isoformat()  # String text parsing test
        else:
            date_of_birth = dob_object              # Native excel object parsing test
            
        ws_students.append([adm_number, student_name, stream, pathway, room, academic_year, term, date_of_birth])

    # ----------------------------------------------------
    # DYNAMIC DOWNLOADS SAVING PATH SYSTEM
    # ----------------------------------------------------
    # Path.home() finds C:\Users\Username on Windows or /Users/Username on Mac
    downloads_folder = Path.home() / "Downloads"
    output_filepath = downloads_folder / "school_import_bulk_test_data.xlsx"
    
    # Save the file using the absolute path string
    wb.save(str(output_filepath))
    
    print(f"✔ Successfully generated bulk template dataset!")
    print(f"📍 SAVED TO: {output_filepath}")
    print(f"  - Subjects   : {len(SUBJECT_TEMPLATES)} rows")
    print(f"  - Teachers   : 40 rows")
    print(f"  - Classrooms : {len(CLASSROOM_MATRIX)} setups mapped")
    print(f"  - Students   : 750 rows")

if __name__ == "__main__":
    generate_bulk_template()