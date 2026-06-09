"""
Management command: import_school_data
=======================================
Imports Teachers, Classrooms (+ Streams), Subjects, and Students
(+ StudentEnrollments) from the official school_import_template.xlsx file.

Usage
-----
    python manage.py import_school_data path/to/data.xlsx
    python manage.py import_school_data path/to/data.xlsx --dry-run

Behaviour
---------
- Import order: Subjects → Teachers → Streams/Classrooms → Students/Enrollments
- Existing records matched by their unique fields are SKIPPED, not duplicated,
  so re-running the command is safe.
- Bad rows are skipped individually and reported at the end.
- --dry-run wraps everything in a rolled-back transaction so you can validate
  the file without touching the database.
"""

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from django.contrib.auth.hashers import make_password
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from school.models import (
    Classroom,
    Stream,
    Student,
    StudentEnrollment,
    Subject,
    Teacher,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell(row, col):
    """Return stripped string value of a cell, or empty string."""
    val = row[col - 1].value
    return str(val).strip() if val is not None else ""


def _require(row, col, label, errors):
    val = _cell(row, col)
    if not val:
        errors.append(f"Missing required field '{label}'")
    return val


def _yn(value: str) -> bool:
    return value.strip().lower() in ("yes", "y", "true", "1")


# ---------------------------------------------------------------------------
# Per-sheet importers
# ---------------------------------------------------------------------------

def import_subjects(ws, stdout, style) -> tuple[int, int, list[str]]:
    created = skipped = 0
    all_errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=3), start=3):
        errors = []
        name       = _require(row, 1, "subject_name", errors)
        department = _require(row, 2, "department",   errors)

        if not name and not department:
            continue  # blank row — end of data

        if errors:
            all_errors.append(f"  Subjects row {row_num}: {'; '.join(errors)}")
            skipped += 1
            continue

        valid_depts = {"MATH", "SCIENCES", "LANGUAGES", "HUMANITIES", "TECHNICALS"}
        if department not in valid_depts:
            all_errors.append(
                f"  Subjects row {row_num}: unknown department '{department}'. "
                f"Must be one of {', '.join(sorted(valid_depts))}."
            )
            skipped += 1
            continue

        _, was_created = Subject.objects.get_or_create(
            name=name, department=department
        )
        created += was_created
        skipped += not was_created

    return created, skipped, all_errors


def import_teachers(ws, stdout, style) -> tuple[int, int, list[str]]:
    created = skipped = 0
    all_errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=3), start=3):
        errors = []
        full_name = _require(row, 1, "full_name", errors)
        username  = _require(row, 2, "username",  errors)
        email     = _cell(row, 3)
        tsc       = _cell(row, 4) or None
        is_hod    = _yn(_cell(row, 5))
        password  = _require(row, 6, "password",  errors)

        if not full_name and not username:
            continue  # blank row

        if errors:
            all_errors.append(f"  Teachers row {row_num}: {'; '.join(errors)}")
            skipped += 1
            continue

        # Clean and normalize username & email matching custom TeacherManager standards
        if hasattr(Teacher, "normalize_username"):
            username = Teacher.normalize_username(username)
        elif hasattr(Teacher.objects, "normalize_username"):
            username = Teacher.objects.normalize_username(username)

        if hasattr(Teacher.objects, "normalize_email"):
            email = Teacher.objects.normalize_email(email) if email else ""

        if Teacher.objects.filter(username=username).exists():
            skipped += 1
            continue

        Teacher.objects.create(
            username=username,
            full_name=full_name,
            email=email,
            tsc_number=tsc,
            is_hod=is_hod,
            is_staff=is_hod,
            password=make_password(password),
        )
        created += 1

    return created, skipped, all_errors


def import_classrooms(ws, stdout, style) -> tuple[int, int, list[str]]:
    """Creates Stream rows as needed, then Classroom rows."""
    created = skipped = 0
    all_errors = []

    valid_pathways = {"STEM", "SOCIAL_SCIENCES", "ARTS_SPORTS", "GENERAL"}

    for row_num, row in enumerate(ws.iter_rows(min_row=3), start=3):
        errors = []
        stream_name    = _require(row, 1, "stream_name",    errors)
        pathway        = _require(row, 2, "pathway",        errors)
        classroom_name = _require(row, 3, "classroom_name", errors)

        if not stream_name and not pathway and not classroom_name:
            continue

        if errors:
            all_errors.append(f"  Classrooms row {row_num}: {'; '.join(errors)}")
            skipped += 1
            continue

        if pathway not in valid_pathways:
            all_errors.append(
                f"  Classrooms row {row_num}: unknown pathway '{pathway}'. "
                f"Must be one of {', '.join(sorted(valid_pathways))}."
            )
            skipped += 1
            continue

        stream, _ = Stream.objects.get_or_create(name=stream_name, pathway=pathway)
        _, was_created = Classroom.objects.get_or_create(
            name=classroom_name, stream=stream
        )
        created += was_created
        skipped += not was_created

    return created, skipped, all_errors


def import_students(ws, stdout, style) -> tuple[int, int, int, list[str]]:
    """
    Creates Student rows and a StudentEnrollment for each.
    Skips enrollment creation if one already exists for the same
    (student, academic_year, term) — allowing the command to be re-run safely.
    """
    students_created = enrollments_created = skipped = 0
    all_errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=3), start=3):
        errors = []
        adm_number     = _require(row, 1, "admission_number", errors)
        name           = _require(row, 2, "name",             errors)
        stream_name    = _require(row, 3, "stream_name",      errors)
        pathway        = _require(row, 4, "pathway",          errors)
        classroom_name = _require(row, 5, "classroom_name",   errors)
        acad_year_raw  = _require(row, 6, "academic_year",    errors)
        term_raw       = _require(row, 7, "term",             errors)

        if not adm_number and not name:
            continue

        if errors:
            all_errors.append(f"  Students row {row_num}: {'; '.join(errors)}")
            skipped += 1
            continue

        # Validate numeric fields
        try:
            academic_year = int(acad_year_raw)
            term = int(term_raw)
            assert 1 <= term <= 3
        except (ValueError, AssertionError):
            all_errors.append(
                f"  Students row {row_num}: academic_year must be a number "
                f"and term must be 1, 2, or 3. Got year='{acad_year_raw}', term='{term_raw}'."
            )
            skipped += 1
            continue

        # Resolve classroom — must already exist (from Classrooms sheet import)
        try:
            stream    = Stream.objects.get(name=stream_name, pathway=pathway)
            classroom = Classroom.objects.get(name=classroom_name, stream=stream)
        except Stream.DoesNotExist:
            all_errors.append(
                f"  Students row {row_num}: stream '{stream_name} / {pathway}' "
                f"not found. Import the Classrooms sheet first, or check spelling."
            )
            skipped += 1
            continue
        except Classroom.DoesNotExist:
            all_errors.append(
                f"  Students row {row_num}: classroom '{classroom_name}' not found "
                f"in stream '{stream_name} / {pathway}'. Check spelling."
            )
            skipped += 1
            continue

        # Parse optional date of birth safely safeguarding openpyxl auto-parsed date fields
        dob = None
        dob_val = row[7].value
        if dob_val is not None:
            if isinstance(dob_val, (date, datetime)):
                dob = dob_val if isinstance(dob_val, date) else dob_val.date()
            else:
                dob_raw = str(dob_val).strip()
                if dob_raw:
                    try:
                        dob = date.fromisoformat(dob_raw)
                    except ValueError:
                        all_errors.append(
                            f"  Students row {row_num}: date_of_birth '{dob_raw}' is not "
                            f"a valid date. Use YYYY-MM-DD. Row will be imported without DOB."
                        )

        # Create or retrieve student
        student, s_created = Student.objects.get_or_create(
            admission_number=adm_number,
            defaults={"name": name, "date_of_birth": dob},
        )
        students_created += s_created

        # Create enrollment if not already present for this year/term
        _, e_created = StudentEnrollment.objects.get_or_create(
            student=student,
            academic_year=academic_year,
            term=term,
            defaults={
                "classroom": classroom,
                "is_active": True,
            },
        )
        enrollments_created += e_created
        if not s_created and not e_created:
            skipped += 1

    return students_created, enrollments_created, skipped, all_errors


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = (
        "Import Teachers, Classrooms, Subjects, and Students from the "
        "official school_import_template.xlsx file."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "file",
            type=str,
            help="Path to the populated school_import_template.xlsx file.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            default=False,
            help=(
                "Validate and preview the import without writing to the database. "
                "All changes are rolled back at the end."
            ),
        )

    def handle(self, *args, **options):
        path = Path(options["file"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            raise CommandError("Only .xlsx / .xlsm files are supported.")

        dry_run = options["dry_run"]
        if dry_run:
            self.stdout.write(self.style.WARNING(
                "\n⚠  DRY RUN — no data will be saved.\n"
            ))

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise CommandError(f"Could not open file: {exc}") from exc

        required_sheets = {"Teachers", "Classrooms", "Subjects", "Students"}
        missing = required_sheets - set(wb.sheetnames)
        if missing:
            raise CommandError(
                f"Missing sheet(s) in workbook: {', '.join(sorted(missing))}. "
                f"Please use the official school_import_template.xlsx."
            )

        all_errors = []

        def _run():
            nonlocal all_errors

            # --- Subjects ---
            self.stdout.write("  Importing subjects...")
            c, s, errs = import_subjects(wb["Subjects"], self.stdout, self.style)
            all_errors.extend(errs)
            self.stdout.write(f"    Created: {c}  |  Skipped/existing: {s}")

            # --- Teachers ---
            self.stdout.write("  Importing teachers...")
            c, s, errs = import_teachers(wb["Teachers"], self.stdout, self.style)
            all_errors.extend(errs)
            self.stdout.write(f"    Created: {c}  |  Skipped/existing: {s}")

            # --- Classrooms (+ Streams) ---
            self.stdout.write("  Importing classrooms and streams...")
            c, s, errs = import_classrooms(wb["Classrooms"], self.stdout, self.style)
            all_errors.extend(errs)
            self.stdout.write(f"    Created: {c}  |  Skipped/existing: {s}")

            # --- Students (+ Enrollments) ---
            self.stdout.write("  Importing students and enrollments...")
            sc, ec, s, errs = import_students(wb["Students"], self.stdout, self.style)
            all_errors.extend(errs)
            self.stdout.write(
                f"    Students created: {sc}  |  "
                f"Enrollments created: {ec}  |  Skipped/existing: {s}"
            )

        self.stdout.write("\nStarting import...\n")

        if dry_run:
            try:
                with transaction.atomic():
                    _run()
                    raise transaction.TransactionManagementError("__dry_run__")
            except transaction.TransactionManagementError as exc:
                if "__dry_run__" not in str(exc):
                    raise
        else:
            with transaction.atomic():
                _run()

        # --- Summary ---
        self.stdout.write("")
        if all_errors:
            self.stdout.write(self.style.WARNING(
                f"⚠  {len(all_errors)} row(s) had errors and were skipped:\n"
            ))
            for err in all_errors:
                self.stdout.write(self.style.WARNING(err))
        else:
            self.stdout.write(self.style.SUCCESS("✔  No errors found."))

        if dry_run:
            self.stdout.write(self.style.WARNING(
                "\nDry run complete — database unchanged. "
                "Remove --dry-run to perform the actual import."
            ))
        else:
            self.stdout.write(self.style.SUCCESS("\n✔  Import complete."))